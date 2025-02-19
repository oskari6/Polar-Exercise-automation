import redis
import pandas as pd
from openpyxl import load_workbook

redis_client = redis.Redis(host='localhost', port=6379, decode_responses=True)

xlsm_file = "exercise_data.xlsm"
book = load_workbook(xlsm_file, keep_vba=True)

#Collect all existing session_ids from the workbook
workbook_session_ids = set()
for sheet_name in book.sheetnames:
    sheet = book[sheet_name]
    workbook_session_ids.update(row[7] for row in sheet.iter_rows(min_row=2, values_only=True) if row[7])

#Fetch data from Redis, excluding existing session_ids
data = []
for year in [2023, 2024, 2025]:
    exercise_ids = redis_client.lrange(f"exercise:{year}", 0, -1)
    for exercise_id in exercise_ids:
        if exercise_id not in workbook_session_ids:
            row = redis_client.hgetall(f"exercise:session:{exercise_id}")
            row["exercise_id"] = exercise_id
            data.append(row)

#Convert the filtered data to a DataFrame
df = pd.DataFrame(data)
if df.empty:
    print("No new data to append.")
    exit()

df["date"] = pd.to_datetime(df["date"], format="%Y-%m-%d %H:%M:%S")
years = df["date"].dt.year.unique()

#Append the filtered data to the appropriate sheets
for year in years:
    yearly_df = df[df["date"].dt.year == year]
    if yearly_df.empty:
        continue
    sheet_name = str(year)
    sheet = book[sheet_name]
    for _, row in yearly_df.iterrows():
        sheet.append(row.tolist())

book.save(xlsm_file)
print("Data successfully written to the workbook!")
