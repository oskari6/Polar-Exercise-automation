import redis
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

redis_client = redis.Redis(host='localhost', port=6379, decode_responses=True)

year = datetime.now().year
xlsm_file = "exercise_data.xlsm"
book = load_workbook(xlsm_file, keep_vba=True)
sheet = book[str(year)]

#Collect all existing exercise_ids from the workbook
workbook_session_ids = set()
workbook_session_ids.update(row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[1])

#Fetch data from Redis, excluding existing session_ids
data = []
exercise_ids = redis_client.lrange(f"exercise:{year}", 0, -1)
session_keys = sorted(redis_client.scan_iter("exercise:session:*"), key=lambda k: int(k.split(":")[-1]), reverse=True)

for exercise_id in exercise_ids:
    if exercise_id not in workbook_session_ids:
        for session_key in session_keys:
            session_data = redis_client.hgetall(session_key)
            if session_data.get("exercise_id") == exercise_id:
                data.append(session_data)
                break

#Convert the filtered data to a DataFrame
df = pd.DataFrame(data)
if df.empty:
    print("No new data to append.")
    exit()

df = df[["session_id","exercise_id","timestamp","date", "duration", "distance", "hr_avg", "hr_max","temperature"]]
df["distance"] = pd.to_numeric(df["distance"],errors="coerce")
df["temperature"] = pd.to_numeric(df["temperature"],errors="coerce")
df["date"] = pd.to_datetime(df["date"]).dt.strftime('%m/%d/%Y')

last_row = max((i for i, row in enumerate(sheet.iter_rows(values_only=True), 1) if any(row[:3])), default=1)

#Append the filtered data
for i, row in enumerate(df.itertuples(index=False), start=last_row+1):
    for col, value in enumerate(row, start=1):  # Start from column 1 (A)
        sheet.cell(row=i, column=col, value=value)

book.save(xlsm_file)
print("Data successfully written to the workbook!")
