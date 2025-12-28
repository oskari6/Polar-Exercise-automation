import os
import sys
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=FutureWarning, message=".*parse_dates.*")
import re
from genericpath import exists
import time

from utils import load_config
from accesslink import AccessLink
import requests
import xml.etree.ElementTree as ET

from weather_api import fetch_weather
import redis

from concurrent.futures import ThreadPoolExecutor
from threading import Lock

import pandas as pd
from openpyxl import load_workbook
import json
import boto3

CONFIG_FILENAME = "config.yml"

config = load_config(CONFIG_FILENAME)

dynamodb = boto3.resource(
    "dynamodb",
    region_name="eu-north-1"
)

accesslink = AccessLink(client_id=config['client_id'],
                        client_secret=config['client_secret'],
                        redirect_url="http://localhost")

redis_client = redis.Redis(
    host=os.getenv("REDIS_HOST", "localhost"),
    port=int(os.getenv("REDIS_PORT", 6379)),
    decode_responses=True
)

def log(msg):
    now = datetime.now()
    print(f"{now:%m/%d/%Y %H:%M:%S}.{now.microsecond // 1000:01d} {msg}")

def parse_iso8601_duration(duration):
    """Convert ISO 8601 duration (PT3766S) to total seconds."""
    match = re.match(r'PT(\d+)(?:\.\d+)?S', duration)
    return int(match.group(1)) if match else 0

def format_duration(seconds):
    """Convert total seconds to h:mm:ss format."""
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{hours}:{minutes:02}:{seconds:02}"

# get coordinates with gpx endpoint
def fetch_location_samples(exercise_id, access_token):
    url = f"https://www.polaraccesslink.com/v3/exercises/{exercise_id}/gpx"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/gpx+xml"
    }
    
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        try:
            root = ET.fromstring(response.text)
            namespace = {'gpx': 'http://www.topografix.com/GPX/1/1'}
            
            trkseg = root.find('.//gpx:trkseg', namespace)
            if trkseg is not None:
                trkpt = trkseg.find('gpx:trkpt', namespace)
                if trkpt is not None:
                    lat = trkpt.attrib.get('lat')
                    lon = trkpt.attrib.get('lon')
                    if lat and lon:
                        lat = format(float(lat), ".4f")
                        lon = format(float(lon), ".4f")
                        return lat, lon
        except ET.ParseError as e:
            log(f"ET Parse Error: {e}")
    else:
        log(f"Failed to fetch GPX data: HTTP {response.status_code} - {response.reason}")
    return None, None

def safe_round(value, ndigits=0, offset=0):
    if value is None:
        return None
    return round(float(value) + offset, ndigits)

def process_exercise(exercise, training_data, access_token, exercise_ids):
    sport = exercise.get("sport", "unknown")
    detailed_sport = exercise.get("detailed_sport_info", "unknown")
    is_treadmill = detailed_sport == "TREADMILL_RUNNING"

    # only running exercises tracked 
    if "RUNNING" in sport:
        exercise_id = exercise.get("id")
        # if already fetched, skip
        if exercise_id in exercise_ids:
            return
        duration_seconds = parse_iso8601_duration(exercise.get('duration'))
        # not shorter than 5min exercises
        if duration_seconds < 300:
            return
        start_time = exercise.get('start_time')
        dt = datetime.strptime(start_time, "%Y-%m-%dT%H:%M:%S")

        distance = None
        temperature = None
        wind_speed = None

        # outside exercises
        if not is_treadmill:
            weather_time = dt.strftime("%Y-%m-%dT%H")
            lat, lon = fetch_location_samples(exercise_id, access_token)

            distance_meters = exercise.get('distance')
            distance = f"{distance_meters / 1000:.2f}" if distance_meters else "0.00"

            if lat is not None and lon is not None:
                weather = fetch_weather(lat, lon, weather_time)
                if weather is not None:
                    temperature = safe_round(weather.get("temperature"), 0, offset=0.5)
                    wind_speed = safe_round(weather.get("wind_speed"), 1)

        heart_rate = exercise.get('heart_rate', {})
        avg_hr = heart_rate.get('average')
        max_hr = heart_rate.get('maximum')

        training_data.append({
            "start_time": dt.strftime("%Y-%m-%d"),
            "duration": format_duration(duration_seconds),
            "distance": distance,
            "hr_avg": avg_hr,
            "hr_max": max_hr,
            "temperature": temperature,
            "wind_speed": wind_speed,
            "timestamp": start_time,
            "exercise_id": exercise_id,
            "treadmill": is_treadmill
        })

def get_additional_data():
    table = dynamodb.Table("TrainingEntries")
    response = table.scan()
    items = response.get("Items", [])

    # Handle pagination
    while "LastEvaluatedKey" in response:
        response = table.scan(
            ExclusiveStartKey=response["LastEvaluatedKey"]
        )
        items.extend(response.get("Items", []))

    return {item["exercise_id"]: item for item in items}

def save_to_redis(training_data):
    additional_data = get_additional_data()

    pipeline = redis_client.pipeline()
    rows = 0

    training_data_sorted = sorted(
        training_data,
        key=lambda d: datetime.strptime(d["timestamp"], "%Y-%m-%dT%H:%M:%S"),
    )

    for data in training_data_sorted:
        rows += 1
        exercise_id = data.get("exercise_id")
        year = datetime.strptime(data['start_time'], "%Y-%m-%d").year
        distance = data.get("distance")
        start_time = data.get("start_time")
        weekday = datetime.strptime(start_time, "%Y-%m-%d").strftime("%a")
        duration = data.get("duration")
        additional_row = additional_data.get(exercise_id)

        if data.get("treadmill"):
            while True:
                try:
                    distance_from_db = additional_row.get("distance")
                    if distance_from_db is not None:
                        distance = distance_from_db
                    else:
                        distance = float(input(f"Enter distance for {start_time} ({weekday}, duration: {duration}): "))
                    break  # Exit loop when valid input is entered
                except ValueError:
                    log("Invalid input. Please enter a valid number.")

        redis_data = {
            "exercise_id": exercise_id,
            "timestamp": data.get("timestamp"),
            "date": start_time,
            "duration": duration,
            "distance": distance,
            "hr_avg": data["hr_avg"],
            "hr_max": data["hr_max"],
            "temperature": data["temperature"],
            "wind_speed": data["wind_speed"],
            "rpe": additional_row.get("rpe"),
            "shoes": additional_row.get("shoes").strip(),
            "notes": additional_row.get("notes").strip(),
        }

        pipeline.rpush(f"exercise:{year}", json.dumps(redis_data))

    pipeline.execute()
    log(f"{rows} row(s) inserted to redis.")

def get_exercise_ids(year):
    redis_key = f"exercise:{year}"
    raw_entries = redis_client.lrange(redis_key, 0, -1)
    return [json.loads(raw).get("exercise_id") for raw in raw_entries if raw]

def parallel_process(exercises, access_token):
    training_data = []
    lock = Lock()
    ids = get_exercise_ids(datetime.now().year)

    def wrapped_process(exercise):
        local_data = []
        process_exercise(exercise, local_data, access_token, ids)
        if local_data:
            with lock:
                training_data.extend(local_data)

    with ThreadPoolExecutor(max_workers=16) as executor:
        executor.map(wrapped_process, exercises)

    return training_data

def insert_data():
    year = datetime.now().year
    xlsm_file = "host_excel/exercise_data.xlsm"
    book = load_workbook(xlsm_file, keep_vba=True)
    sheet = book[str(year)]

    workbook_exercise_ids = set(
        row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]
    )

    redis_key = f"exercise:{year}"
    if not redis_client.exists(redis_key):
        log(f"Warning: No Redis list found for {year}. Nothing to insert.")
        return

    raw_entries = redis_client.lrange(redis_key, 0, -1)
    data = []
    for raw in raw_entries:
        entry = json.loads(raw)
        if entry.get("exercise_id") not in workbook_exercise_ids:
            data.append(entry)

    #Convert the filtered data to a DataFrame
    df = pd.DataFrame(data)
    if df.empty:
        log("No new data to append.")
        exit(2)

    df = df[["exercise_id",
        "timestamp",
        "date",
        "duration",
        "distance",
        "hr_avg",
        "hr_max",
        "temperature",
        "wind_speed",
        "rpe",
        "shoes",
        "notes"]]
    df["distance"] = pd.to_numeric(df["distance"],errors="coerce")
    df["temperature"] = pd.to_numeric(df["temperature"],errors="coerce")
    df["wind_speed"] = pd.to_numeric(df["wind_speed"],errors="coerce")
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    df = df.sort_values(by="date", ascending=True).reset_index(drop=True)

    EXCEL_COLUMN_MAP = {
        "exercise_id": 1,   # A (if present, else remove)
        "timestamp": 2,     # or shift if exercise_id not shown
        "date": 3,
        "duration": 4,
        "distance": 5,
        "hr_avg": 6,
        "hr_max": 7,
        "temperature": 8,
        "wind_speed": 9,
        # column 10 = pace avg (SKIP)
        "rpe": 11,
        "shoes": 12,
        "notes": 13,
    }

    last_row = 1
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            last_row = i

    #Append the filtered data
    for i, record in enumerate(df.to_dict("records"), start=last_row + 1):
        for field, col in EXCEL_COLUMN_MAP.items():
            value = record.get(field)
            cell = sheet.cell(row=i, column=col, value=value)

            if field == "date" and value is not None:
                cell.number_format = "DD-MMM"

    book.save(xlsm_file)

def clear_table():
    table = dynamodb.Table("TrainingEntries")

    response = table.scan(ProjectionExpression="id")

    with table.batch_writer() as batch:
        for item in response["Items"]:
            batch.delete_item(
                Key={"id": item["id"]}
            )

    while "LastEvaluatedKey" in response:
        response = table.scan(
            ProjectionExpression="id",
            ExclusiveStartKey=response["LastEvaluatedKey"],
        )
        with table.batch_writer() as batch:
            for item in response["Items"]:
                batch.delete_item(
                    Key={"id": item["id"]})

# main fetch
def fetch_data(fetch_exercises_only = False):
    training_data = []
    access_token = config["access_token"]
    exercises = accesslink.get_exercises(access_token=access_token)
    if fetch_exercises_only:
        return exercises
    training_data = parallel_process(exercises, access_token)

    if training_data:
        save_to_redis(training_data)
        log("Inserting data...")
        insert_data()
        log("Data successfully written to the workbook")
        clear_table()
        log("DynamoDB table cleared")
        return sys.exit(0)
    return sys.exit(2)

if __name__ == "__main__":
    log("Fetching data...")
    fetch_data()

# docker compose run --rm app python fetch_data.py